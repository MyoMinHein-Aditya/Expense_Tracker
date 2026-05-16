from flask import Flask, request, jsonify, render_template, send_file
import sqlite3
from datetime import datetime
import io

app = Flask(__name__)
DB = "expenses.db"

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS balance (
            id INTEGER PRIMARY KEY, amount REAL NOT NULL DEFAULT 0)""")
        conn.execute("""CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            description TEXT NOT NULL, amount REAL NOT NULL,
            type TEXT NOT NULL, created_at TEXT NOT NULL)""")
        if conn.execute("SELECT COUNT(*) as c FROM balance").fetchone()["c"] == 0:
            conn.execute("INSERT INTO balance (amount) VALUES (0)")
        conn.commit()

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/balance")
def get_balance():
    with get_db() as conn:
        row = conn.execute("SELECT amount FROM balance WHERE id=1").fetchone()
        return jsonify({"balance": row["amount"]})

@app.route("/api/add-money", methods=["POST"])
def add_money():
    data = request.json
    amount = float(data.get("amount", 0))
    note = data.get("note", "Added money")
    if amount <= 0:
        return jsonify({"error": "Invalid amount"}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as conn:
        conn.execute("UPDATE balance SET amount = amount + ? WHERE id=1", (amount,))
        conn.execute("INSERT INTO transactions (description, amount, type, created_at) VALUES (?, ?, 'credit', ?)", (note, amount, now))
        conn.commit()
        row = conn.execute("SELECT amount FROM balance WHERE id=1").fetchone()
    return jsonify({"balance": row["amount"]})

@app.route("/api/add-expense", methods=["POST"])
def add_expense():
    data = request.json
    description = data.get("description", "").strip()
    amount = float(data.get("amount", 0))
    if not description or amount <= 0:
        return jsonify({"error": "Invalid data"}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as conn:
        bal = conn.execute("SELECT amount FROM balance WHERE id=1").fetchone()["amount"]
        if amount > bal:
            return jsonify({"error": "Insufficient balance"}), 400
        conn.execute("UPDATE balance SET amount = amount - ? WHERE id=1", (amount,))
        conn.execute("INSERT INTO transactions (description, amount, type, created_at) VALUES (?, ?, 'debit', ?)", (description, amount, now))
        conn.commit()
        row = conn.execute("SELECT amount FROM balance WHERE id=1").fetchone()
    return jsonify({"balance": row["amount"]})

@app.route("/api/transactions")
def get_transactions():
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    with get_db() as conn:
        rows = conn.execute("""SELECT id, description, amount, type, created_at
            FROM transactions WHERE strftime('%Y-%m', created_at) = ?
            ORDER BY created_at DESC""", (month,)).fetchall()
        txns = [dict(r) for r in rows]
        total_spent = sum(t["amount"] for t in txns if t["type"] == "debit")
        total_added = sum(t["amount"] for t in txns if t["type"] == "credit")
    return jsonify({"transactions": txns, "total_spent": total_spent, "total_added": total_added})

@app.route("/api/delete/<int:txn_id>", methods=["DELETE"])
def delete_transaction(txn_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM transactions WHERE id=?", (txn_id,)).fetchone()
        if not row:
            return jsonify({"error": "Not found"}), 404
        if row["type"] == "debit":
            conn.execute("UPDATE balance SET amount = amount + ? WHERE id=1", (row["amount"],))
        else:
            conn.execute("UPDATE balance SET amount = amount - ? WHERE id=1", (row["amount"],))
        conn.execute("DELETE FROM transactions WHERE id=?", (txn_id,))
        conn.commit()
        bal = conn.execute("SELECT amount FROM balance WHERE id=1").fetchone()
    return jsonify({"balance": bal["amount"]})

# ── CHARTS ──────────────────────────────────────────────────────────────────

def chart_style(fig, ax):
    """Apply dark theme to any chart."""
    BG = "#0e0e0e"
    fig.patch.set_facecolor(BG)
    ax.set_facecolor("#1a1a1a")
    ax.tick_params(colors="#888888", labelsize=8)
    ax.spines[:].set_color("#2e2e2e")
    for spine in ax.spines.values():
        spine.set_linewidth(0.5)

@app.route("/api/chart/daily")
def chart_daily():
    """Bar chart: daily spending for current month."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import calendar

    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    year, mon = map(int, month.split("-"))
    days_in_month = calendar.monthrange(year, mon)[1]

    with get_db() as conn:
        rows = conn.execute("""
            SELECT CAST(strftime('%d', created_at) AS INTEGER) as day,
                   SUM(CASE WHEN type='debit' THEN amount ELSE 0 END) as spent
            FROM transactions
            WHERE strftime('%Y-%m', created_at) = ?
            GROUP BY day ORDER BY day
        """, (month,)).fetchall()

    data = {r["day"]: r["spent"] for r in rows}
    days = list(range(1, days_in_month + 1))
    values = [data.get(d, 0) for d in days]

    fig, ax = plt.subplots(figsize=(7, 2.6))
    chart_style(fig, ax)

    colors = ["#c8f04a" if v == max(values) and v > 0 else "#ff5f5f" if v > 0 else "#2e2e2e" for v in values]
    bars = ax.bar(days, values, color=colors, width=0.7, zorder=3)
    ax.set_xlim(0.5, days_in_month + 0.5)
    ax.set_xlabel("Day of Month", color="#888", fontsize=8, labelpad=6)
    ax.set_ylabel("₹ Spent", color="#888", fontsize=8, labelpad=6)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"₹{x/1000:.0f}k" if x >= 1000 else f"₹{x:.0f}"))
    ax.grid(axis="y", color="#2e2e2e", linewidth=0.5, zorder=0)
    ax.set_title("Daily Spending — " + datetime.strptime(month, "%Y-%m").strftime("%B %Y"),
                 color="#f0f0f0", fontsize=10, pad=10, fontweight="bold")
    plt.tight_layout(pad=1.2)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor=fig.get_facecolor())
    buf.seek(0)
    plt.close(fig)
    return send_file(buf, mimetype="image/png")

@app.route("/api/chart/monthly")
def chart_monthly():
    """Grouped bar chart: monthly income vs spending (last 6 months)."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    with get_db() as conn:
        rows = conn.execute("""
            SELECT strftime('%Y-%m', created_at) as month, type, SUM(amount) as total
            FROM transactions GROUP BY month, type ORDER BY month ASC
        """).fetchall()

    months_raw = {}
    for r in rows:
        m = r["month"]
        months_raw.setdefault(m, {"spent": 0, "added": 0})
        months_raw[m]["debit" == r["type"] and "spent" or "added"] = r["total"]

    # fix: proper key assignment
    months_data = {}
    for r in rows:
        m = r["month"]
        months_data.setdefault(m, {"spent": 0, "added": 0})
        if r["type"] == "debit":
            months_data[m]["spent"] = r["total"]
        else:
            months_data[m]["added"] = r["total"]

    # Last 6 months
    sorted_months = sorted(months_data.keys())[-6:]
    labels = [datetime.strptime(m, "%Y-%m").strftime("%b '%y") for m in sorted_months]
    spent  = [months_data[m]["spent"]  for m in sorted_months]
    added  = [months_data[m]["added"]  for m in sorted_months]

    x = np.arange(len(labels))
    w = 0.35

    fig, ax = plt.subplots(figsize=(7, 2.6))
    chart_style(fig, ax)

    ax.bar(x - w/2, added, w, label="Income",  color="#4affa0", alpha=0.9, zorder=3)
    ax.bar(x + w/2, spent, w, label="Expenses", color="#ff5f5f", alpha=0.9, zorder=3)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, color="#888", fontsize=8)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"₹{v/1000:.0f}k" if v >= 1000 else f"₹{v:.0f}"))
    ax.grid(axis="y", color="#2e2e2e", linewidth=0.5, zorder=0)
    ax.legend(fontsize=8, facecolor="#1a1a1a", edgecolor="#2e2e2e",
              labelcolor="#f0f0f0", loc="upper left")
    ax.set_title("Monthly Overview (Last 6 Months)", color="#f0f0f0", fontsize=10, pad=10, fontweight="bold")
    plt.tight_layout(pad=1.2)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor=fig.get_facecolor())
    buf.seek(0)
    plt.close(fig)
    return send_file(buf, mimetype="image/png")

@app.route("/api/chart/breakdown")
def chart_breakdown():
    """Donut chart: top expense categories this month."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    with get_db() as conn:
        rows = conn.execute("""
            SELECT description, SUM(amount) as total
            FROM transactions
            WHERE type='debit' AND strftime('%Y-%m', created_at) = ?
            GROUP BY description ORDER BY total DESC LIMIT 7
        """, (month,)).fetchall()

    if not rows:
        fig, ax = plt.subplots(figsize=(4, 2.6))
        chart_style(fig, ax)
        ax.text(0.5, 0.5, "No expenses yet", ha="center", va="center",
                color="#888", fontsize=10, transform=ax.transAxes)
        ax.axis("off")
        ax.set_title("Expense Breakdown", color="#f0f0f0", fontsize=10, pad=10, fontweight="bold")
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor=fig.get_facecolor())
        buf.seek(0); plt.close(fig)
        return send_file(buf, mimetype="image/png")

    labels = [r["description"] for r in rows]
    values = [r["total"] for r in rows]
    palette = ["#c8f04a","#ff5f5f","#4affa0","#ff9f40","#36a2eb","#9966ff","#ff6384"]

    fig, ax = plt.subplots(figsize=(4, 2.6))
    chart_style(fig, ax)
    wedges, texts, autotexts = ax.pie(
        values, labels=None, autopct="%1.0f%%",
        colors=palette[:len(values)], startangle=90,
        wedgeprops=dict(width=0.55, edgecolor="#0e0e0e", linewidth=1.5),
        pctdistance=0.75
    )
    for at in autotexts:
        at.set_color("#0e0e0e"); at.set_fontsize(7); at.set_fontweight("bold")

    ax.legend(wedges, [f"{l[:14]}" for l in labels],
              fontsize=7, facecolor="#1a1a1a", edgecolor="#2e2e2e",
              labelcolor="#f0f0f0", loc="center left",
              bbox_to_anchor=(0.85, 0.5), framealpha=0.9)
    ax.set_title("Expense Breakdown", color="#f0f0f0", fontsize=10, pad=10, fontweight="bold")
    plt.tight_layout(pad=0.5)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor=fig.get_facecolor())
    buf.seek(0); plt.close(fig)
    return send_file(buf, mimetype="image/png")

# ── EXPORT ──────────────────────────────────────────────────────────────────

@app.route("/api/export")
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "Run: pip install openpyxl"}), 500

    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    with get_db() as conn:
        rows = conn.execute("""SELECT description, amount, type, created_at
            FROM transactions WHERE strftime('%Y-%m', created_at) = ?
            ORDER BY created_at ASC""", (month,)).fetchall()
        txns = [dict(r) for r in rows]
        bal = conn.execute("SELECT amount FROM balance WHERE id=1").fetchone()["amount"]

    wb = openpyxl.Workbook()
    ws = wb.active
    month_label = datetime.strptime(month, "%Y-%m").strftime("%B %Y")
    ws.title = month_label

    dark  = PatternFill("solid", fgColor="0E0E0E")
    dark2 = PatternFill("solid", fgColor="1A1A1A")
    rowa  = PatternFill("solid", fgColor="111111")
    rowb  = PatternFill("solid", fgColor="161616")
    facc  = Font(name="Calibri", bold=True, color="C8F04A", size=14)
    fhdr  = Font(name="Calibri", bold=True, color="FFFFFF",  size=11)
    fred  = Font(name="Calibri", bold=True, color="FF5F5F",  size=11)
    fgrn  = Font(name="Calibri", bold=True, color="4AFFA0",  size=11)
    fyel  = Font(name="Calibri", bold=True, color="C8F04A",  size=11)
    fmut  = Font(name="Calibri", color="888888", size=10)
    fnor  = Font(name="Calibri", color="F0F0F0", size=11)
    thin  = Border(bottom=Side(style="thin",   color="2E2E2E"))
    thick = Border(bottom=Side(style="medium", color="C8F04A"))
    ctr   = Alignment(horizontal="left", vertical="center")

    def sc(cell, val, font, fill, align=ctr):
        cell.value=val; cell.font=font; cell.fill=fill; cell.alignment=align

    total_spent = sum(t["amount"] for t in txns if t["type"] == "debit")
    total_added = sum(t["amount"] for t in txns if t["type"] == "credit")

    ws.merge_cells("A1:D1"); sc(ws["A1"], f"PaisaTrack — {month_label}", facc, dark); ws.row_dimensions[1].height=32
    ws.merge_cells("A2:B2"); sc(ws["A2"], f"Total Added:  ₹{total_added:,.2f}", fgrn, dark)
    ws.merge_cells("C2:D2"); sc(ws["C2"], f"Total Spent:  ₹{total_spent:,.2f}", fred, dark); ws.row_dimensions[2].height=22
    ws.merge_cells("A3:D3"); sc(ws["A3"], f"Current Balance:  ₹{bal:,.2f}", fyel, dark); ws.row_dimensions[3].height=22
    for c in "ABCD": ws[f"{c}4"].fill=dark
    ws.row_dimensions[4].height=8

    for i,h in enumerate(["Date & Time","Description","Type","Amount (₹)"],1):
        cell=ws.cell(row=5,column=i,value=h)
        cell.font=fhdr; cell.fill=dark2; cell.alignment=ctr; cell.border=thick
    ws.row_dimensions[5].height=22

    for ri, t in enumerate(txns, 6):
        dt = datetime.strptime(t["created_at"], "%Y-%m-%d %H:%M:%S")
        debit = t["type"]=="debit"
        fill = rowa if ri%2==0 else rowb
        vals = [dt.strftime("%d %b, %I:%M %p"), t["description"],
                "Expense" if debit else "Added",
                f"{'−' if debit else '+'} ₹{t['amount']:,.2f}"]
        fonts = [fmut, fnor, Font(name="Calibri",color="FF5F5F" if debit else "4AFFA0",size=10),
                 fred if debit else fgrn]
        for ci,(v,f) in enumerate(zip(vals,fonts),1):
            cell=ws.cell(row=ri,column=ci,value=v)
            cell.font=f; cell.fill=fill; cell.border=thin; cell.alignment=ctr
        ws.row_dimensions[ri].height=20

    ws.column_dimensions["A"].width=22
    ws.column_dimensions["B"].width=36
    ws.column_dimensions["C"].width=12
    ws.column_dimensions["D"].width=20
    ws.sheet_view.showGridLines=False

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"PaisaTrack_{month}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    init_db()
    print("\n🚀 Expense Tracker running at http://localhost:5000\n")
    app.run(debug=True, port=5000)
